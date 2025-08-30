import openpyxl
from openpyxl import Workbook, cell

wb = openpyxl.Workbook("Example.xlsx")
wb.create_sheet("Sheet1")
ws = wb["Sheet1"]
ws = wb.active  
#ws.title = "MySheet"
ws["A1"] = "Hello"
ws["B1"] = "World"
wb.save("Example.xlsx")
wb.close()